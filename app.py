import os
import sqlite3
import secrets
from datetime import datetime
from flask import Flask, request, jsonify, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('SESSION_SECRET', secrets.token_hex(32))

DATABASE = 'contract_review.db'

def get_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    return db

def init_db():
    db = get_db()
    db.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            name TEXT NOT NULL,
            department TEXT NOT NULL,
            is_admin BOOLEAN NOT NULL DEFAULT 0,
            lead_form_access BOOLEAN NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    try:
        db.execute('SELECT lead_form_access FROM users LIMIT 1')
    except:
        db.execute('ALTER TABLE users ADD COLUMN lead_form_access BOOLEAN NOT NULL DEFAULT 0')
        db.commit()
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS pos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer TEXT NOT NULL,
            bid TEXT NOT NULL,
            po TEXT NOT NULL,
            cr TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS cr_forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_key TEXT UNIQUE NOT NULL,
            customer TEXT,
            bid TEXT,
            po TEXT,
            cr TEXT,
            record_no TEXT,
            record_date TEXT,
            last_modified_by TEXT,
            last_modified_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    try:
        db.execute('SELECT record_no FROM cr_forms LIMIT 1')
    except:
        db.execute('ALTER TABLE cr_forms ADD COLUMN record_no TEXT')
        db.execute('ALTER TABLE cr_forms ADD COLUMN record_date TEXT')
        db.commit()
    
    try:
        db.execute('SELECT amendment_details FROM cr_forms LIMIT 1')
    except:
        db.execute('ALTER TABLE cr_forms ADD COLUMN amendment_details TEXT')
        db.commit()
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS cr_form_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cr_form_id INTEGER NOT NULL,
            item_no TEXT NOT NULL,
            part_number TEXT,
            part_description TEXT,
            rev TEXT,
            qty TEXT,
            cycles TEXT,
            remarks TEXT,
            FOREIGN KEY (cr_form_id) REFERENCES cr_forms(id) ON DELETE CASCADE
        )
    ''')
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS ped_forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_key TEXT UNIQUE NOT NULL,
            customer TEXT,
            bid TEXT,
            po TEXT,
            cr TEXT,
            record_no TEXT,
            record_date TEXT,
            amendment_details TEXT,
            last_modified_by TEXT,
            last_modified_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS ped_form_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ped_form_id INTEGER NOT NULL,
            item_no TEXT NOT NULL,
            part_number TEXT,
            part_description TEXT,
            rev TEXT,
            qty TEXT,
            ped_cycles TEXT,
            notes TEXT,
            remarks TEXT,
            FOREIGN KEY (ped_form_id) REFERENCES ped_forms(id) ON DELETE CASCADE
        )
    ''')
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS lead_forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_key TEXT UNIQUE NOT NULL,
            customer TEXT,
            bid TEXT,
            po TEXT,
            cr TEXT,
            record_no TEXT,
            record_date TEXT,
            last_modified_by TEXT,
            last_modified_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    db.execute('''
        CREATE TABLE IF NOT EXISTS lead_form_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_form_id INTEGER NOT NULL,
            item_no TEXT NOT NULL,
            part_number TEXT,
            part_description TEXT,
            rev TEXT,
            qty TEXT,
            customer_required_date TEXT,
            standard_lead_time TEXT,
            gtn_agreed_date TEXT,
            remarks TEXT,
            FOREIGN KEY (lead_form_id) REFERENCES lead_forms(id) ON DELETE CASCADE
        )
    ''')
    
    cursor = db.execute('SELECT COUNT(*) as count FROM users WHERE username = ?', ('admin',))
    if cursor.fetchone()['count'] == 0:
        db.execute(
            'INSERT INTO users (username, password_hash, name, department, is_admin) VALUES (?, ?, ?, ?, ?)',
            ('admin', generate_password_hash('admin'), 'IT Administrator', 'it', 1)
        )
    
    db.commit()
    db.close()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        db = get_db()
        user = db.execute('SELECT is_admin FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        db.close()
        if not user or not user['is_admin']:
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    return send_from_directory('static', 'login.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
    db.close()
    
    if user and check_password_hash(user['password_hash'], password):
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['user_department'] = user['department']
        session['user_name'] = user['name']
        session['user_is_admin'] = bool(user['is_admin'])
        
        return jsonify({
            'success': True,
            'user': {
                'username': user['username'],
                'name': user['name'],
                'department': user['department'],
                'isAdmin': bool(user['is_admin'])
            }
        })
    
    return jsonify({'error': 'Invalid username or password'}), 401

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/session', methods=['GET'])
def get_session():
    if 'user_id' in session:
        return jsonify({
            'loggedIn': True,
            'user': {
                'username': session.get('username'),
                'name': session.get('user_name'),
                'department': session.get('user_department'),
                'isAdmin': session.get('user_is_admin', False)
            }
        })
    return jsonify({'loggedIn': False})

@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    db = get_db()
    users = db.execute('SELECT id, username, name, department, is_admin, lead_form_access FROM users ORDER BY id').fetchall()
    db.close()
    
    return jsonify([{
        'id': user['id'],
        'username': user['username'],
        'name': user['name'],
        'department': user['department'],
        'isAdmin': bool(user['is_admin']),
        'leadFormAccess': bool(user['lead_form_access'])
    } for user in users])

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    name = data.get('name', '').strip()
    department = data.get('department', '').strip()
    is_admin = data.get('isAdmin', False)
    lead_form_access = data.get('leadFormAccess', False)
    
    if not all([username, password, name, department]):
        return jsonify({'error': 'All fields required'}), 400
    
    db = get_db()
    
    existing = db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
    if existing:
        db.close()
        return jsonify({'error': 'Username already exists'}), 400
    
    try:
        db.execute(
            'INSERT INTO users (username, password_hash, name, department, is_admin, lead_form_access) VALUES (?, ?, ?, ?, ?, ?)',
            (username, generate_password_hash(password), name, department, 1 if is_admin else 0, 1 if lead_form_access else 0)
        )
        db.commit()
        user_id = db.execute('SELECT last_insert_rowid() as id').fetchone()['id']
        db.close()
        
        return jsonify({
            'success': True,
            'user': {
                'id': user_id,
                'username': username,
                'name': name,
                'department': department,
                'isAdmin': is_admin,
                'leadFormAccess': lead_form_access
            }
        })
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user(user_id):
    data = request.get_json()
    name = data.get('name', '').strip()
    department = data.get('department', '').strip()
    is_admin = data.get('isAdmin', False)
    lead_form_access = data.get('leadFormAccess', False)
    password = data.get('password', '').strip()
    
    if not all([name, department]):
        return jsonify({'error': 'Name and department required'}), 400
    
    db = get_db()
    
    user = db.execute('SELECT username, is_admin FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'User not found'}), 404
    
    if user['username'] == 'admin' and user['is_admin'] and not is_admin:
        admin_count = db.execute('SELECT COUNT(*) as count FROM users WHERE is_admin = 1').fetchone()['count']
        if admin_count <= 1:
            db.close()
            return jsonify({'error': 'Cannot remove admin status from last admin user'}), 400
    
    try:
        if password:
            db.execute(
                'UPDATE users SET name = ?, department = ?, is_admin = ?, lead_form_access = ?, password_hash = ? WHERE id = ?',
                (name, department, 1 if is_admin else 0, 1 if lead_form_access else 0, generate_password_hash(password), user_id)
            )
        else:
            db.execute(
                'UPDATE users SET name = ?, department = ?, is_admin = ?, lead_form_access = ? WHERE id = ?',
                (name, department, 1 if is_admin else 0, 1 if lead_form_access else 0, user_id)
            )
        db.commit()
        db.close()
        
        return jsonify({
            'success': True,
            'user': {
                'id': user_id,
                'username': user['username'],
                'name': name,
                'department': department,
                'isAdmin': is_admin,
                'leadFormAccess': lead_form_access
            }
        })
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    db = get_db()
    
    user = db.execute('SELECT username, is_admin FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'User not found'}), 404
    
    if user['username'] == 'admin' and user['is_admin']:
        db.close()
        return jsonify({'error': 'Cannot delete default admin user'}), 400
    
    admin_count = db.execute('SELECT COUNT(*) as count FROM users WHERE is_admin = 1').fetchone()['count']
    if user['is_admin'] and admin_count <= 1:
        db.close()
        return jsonify({'error': 'Cannot delete last admin user'}), 400
    
    db.execute('DELETE FROM users WHERE id = ?', (user_id,))
    db.commit()
    db.close()
    
    return jsonify({'success': True})

@app.route('/api/pos', methods=['GET'])
@login_required
def get_pos():
    db = get_db()
    pos_list = db.execute('SELECT * FROM pos ORDER BY id DESC').fetchall()
    db.close()
    
    return jsonify([{
        'id': po['id'],
        'customer': po['customer'],
        'bid': po['bid'],
        'po': po['po'],
        'cr': po['cr']
    } for po in pos_list])

@app.route('/api/pos', methods=['POST'])
@admin_required
def create_po():
    data = request.get_json()
    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    
    if not all([customer, bid, po, cr]):
        return jsonify({'error': 'All fields required'}), 400
    
    db = get_db()
    try:
        db.execute(
            'INSERT INTO pos (customer, bid, po, cr) VALUES (?, ?, ?, ?)',
            (customer, bid, po, cr)
        )
        db.commit()
        po_id = db.execute('SELECT last_insert_rowid() as id').fetchone()['id']
        db.close()
        
        return jsonify({
            'success': True,
            'po': {
                'id': po_id,
                'customer': customer,
                'bid': bid,
                'po': po,
                'cr': cr
            }
        })
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/pos/<int:po_id>', methods=['PUT'])
@admin_required
def update_po(po_id):
    data = request.get_json()
    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    
    if not all([customer, bid, po, cr]):
        return jsonify({'error': 'All fields required'}), 400
    
    db = get_db()
    db.execute(
        'UPDATE pos SET customer = ?, bid = ?, po = ?, cr = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
        (customer, bid, po, cr, po_id)
    )
    db.commit()
    db.close()
    
    return jsonify({
        'success': True,
        'po': {
            'id': po_id,
            'customer': customer,
            'bid': bid,
            'po': po,
            'cr': cr
        }
    })

@app.route('/api/pos/<int:po_id>', methods=['DELETE'])
@admin_required
def delete_po(po_id):
    db = get_db()
    db.execute('DELETE FROM pos WHERE id = ?', (po_id,))
    db.commit()
    db.close()
    
    return jsonify({'success': True})

@app.route('/api/backup', methods=['GET'])
@admin_required
def backup_data():
    db = get_db()
    users = db.execute('SELECT id, username, password_hash, name, department, is_admin FROM users').fetchall()
    pos_list = db.execute('SELECT * FROM pos').fetchall()
    db.close()
    
    return jsonify({
        'meta': {
            'app': 'GTN-ContractReview',
            'version': '1.0',
            'exportedAt': datetime.utcnow().isoformat(),
            'by': session.get('username', '')
        },
        'users': [{
            'username': u['username'],
            'password_hash': u['password_hash'],
            'name': u['name'],
            'department': u['department'],
            'isAdmin': bool(u['is_admin'])
        } for u in users],
        'pos': [{
            'customer': p['customer'],
            'bid': p['bid'],
            'po': p['po'],
            'cr': p['cr']
        } for p in pos_list]
    })

@app.route('/api/restore', methods=['POST'])
@admin_required
def restore_data():
    data = request.get_json()
    
    if not data or not isinstance(data, dict):
        return jsonify({'error': 'Invalid data format'}), 400
    
    if not data.get('users') or not isinstance(data['users'], list):
        return jsonify({'error': 'Missing or invalid users array'}), 400
    
    if not data.get('pos') or not isinstance(data['pos'], list):
        return jsonify({'error': 'Missing or invalid pos array'}), 400
    
    for user in data['users']:
        if not all(key in user for key in ['username', 'password_hash', 'name', 'department']):
            return jsonify({'error': 'Invalid user data: missing required fields'}), 400
        if not user.get('password_hash') or not user['password_hash'].startswith(('pbkdf2:', 'scrypt:', 'bcrypt:')):
            return jsonify({'error': 'Invalid user data: password_hash must be a valid hash'}), 400
    
    has_admin = any(u.get('isAdmin') for u in data['users'])
    if not has_admin:
        return jsonify({'error': 'Backup must contain at least one admin user'}), 400
    
    for po in data['pos']:
        if not all(key in po for key in ['customer', 'bid', 'po', 'cr']):
            return jsonify({'error': 'Invalid PO data: missing required fields'}), 400
    
    db = get_db()
    try:
        db.execute('BEGIN TRANSACTION')
        
        db.execute('DELETE FROM users')
        db.execute('DELETE FROM pos')
        
        for user in data['users']:
            db.execute(
                'INSERT INTO users (username, password_hash, name, department, is_admin) VALUES (?, ?, ?, ?, ?)',
                (user['username'], user['password_hash'], 
                 user['name'], user['department'], 1 if user.get('isAdmin') else 0)
            )
        
        for po in data['pos']:
            db.execute(
                'INSERT INTO pos (customer, bid, po, cr) VALUES (?, ?, ?, ?)',
                (po['customer'], po['bid'], po['po'], po['cr'])
            )
        
        db.commit()
        db.close()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cr-form/save', methods=['POST'])
@login_required
def save_cr_form():
    import json
    data = request.get_json()
    
    po_key = data.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400
    
    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    record_no = data.get('recordNo', '').strip()
    record_date = data.get('recordDate', '').strip()
    rows = data.get('rows', [])
    
    username = session.get('username', 'unknown')
    is_admin = session.get('user_is_admin', False)
    
    db = get_db()
    try:
        db.execute('BEGIN TRANSACTION')
        
        cursor = db.execute('SELECT id, amendment_details FROM cr_forms WHERE po_key = ?', (po_key,))
        form = cursor.fetchone()
        
        if is_admin:
            amendment_details = data.get('amendmentDetails', '').strip()
        else:
            amendment_details = form['amendment_details'] if form else ''
        
        if form:
            form_id = form['id']
            db.execute('''
                UPDATE cr_forms 
                SET customer = ?, bid = ?, po = ?, cr = ?, record_no = ?, record_date = ?, amendment_details = ?,
                    last_modified_by = ?, last_modified_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (customer, bid, po, cr, record_no, record_date, amendment_details, username, form_id))
            
            db.execute('DELETE FROM cr_form_rows WHERE cr_form_id = ?', (form_id,))
        else:
            db.execute('''
                INSERT INTO cr_forms (po_key, customer, bid, po, cr, record_no, record_date, amendment_details, last_modified_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (po_key, customer, bid, po, cr, record_no, record_date, amendment_details, username))
            form_id = db.execute('SELECT last_insert_rowid() as id').fetchone()['id']
        
        for row in rows:
            item_no = row.get('key', '')
            if not item_no:
                continue
            
            cycles_json = json.dumps(row.get('cycles', []))
            
            db.execute('''
                INSERT INTO cr_form_rows 
                (cr_form_id, item_no, part_number, part_description, rev, qty, cycles, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                form_id,
                item_no,
                row.get('part', ''),
                row.get('desc', ''),
                row.get('rev', ''),
                row.get('qty', ''),
                cycles_json,
                row.get('remarks', '')
            ))
        
        db.commit()
        db.close()
        
        return jsonify({
            'success': True,
            'lastModifiedBy': username,
            'lastModifiedAt': datetime.utcnow().isoformat()
        })
    except Exception as e:
        db.rollback()
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cr-form/load', methods=['GET'])
@login_required
def load_cr_form():
    import json
    po_key = request.args.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400
    
    db = get_db()
    try:
        cursor = db.execute('''
            SELECT id, customer, bid, po, cr, record_no, record_date, amendment_details, last_modified_by, last_modified_at
            FROM cr_forms
            WHERE po_key = ?
        ''', (po_key,))
        form = cursor.fetchone()
        
        if not form:
            db.close()
            return jsonify({'exists': False})
        
        form_id = form['id']
        
        rows_cursor = db.execute('''
            SELECT item_no, part_number, part_description, rev, qty, cycles, remarks
            FROM cr_form_rows
            WHERE cr_form_id = ?
            ORDER BY id
        ''', (form_id,))
        
        rows = []
        for row in rows_cursor.fetchall():
            cycles = json.loads(row['cycles']) if row['cycles'] else []
            rows.append({
                'key': row['item_no'],
                'part': row['part_number'] or '',
                'desc': row['part_description'] or '',
                'rev': row['rev'] or '',
                'qty': row['qty'] or '',
                'cycles': cycles,
                'remarks': row['remarks'] or ''
            })
        
        db.close()
        
        return jsonify({
            'exists': True,
            'customer': form['customer'] or '',
            'bid': form['bid'] or '',
            'po': form['po'] or '',
            'cr': form['cr'] or '',
            'recordNo': form['record_no'] or '',
            'recordDate': form['record_date'] or '',
            'amendmentDetails': form['amendment_details'] or '',
            'rows': rows,
            'lastModifiedBy': form['last_modified_by'] or '',
            'lastModifiedAt': form['last_modified_at'] or ''
        })
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ped-form/save', methods=['POST'])
@login_required
def save_ped_form():
    import json
    data = request.get_json()
    
    po_key = data.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400
    
    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    record_no = data.get('recordNo', '').strip()
    record_date = data.get('recordDate', '').strip()
    rows = data.get('rows', [])
    
    username = session.get('username', 'unknown')
    is_admin = session.get('user_is_admin', False)
    
    db = get_db()
    try:
        db.execute('BEGIN TRANSACTION')
        
        cursor = db.execute('SELECT id, amendment_details FROM ped_forms WHERE po_key = ?', (po_key,))
        form = cursor.fetchone()
        
        if is_admin:
            amendment_details = data.get('amendmentDetails', '').strip()
        else:
            amendment_details = form['amendment_details'] if form else ''
        
        if form:
            form_id = form['id']
            db.execute('''
                UPDATE ped_forms 
                SET customer = ?, bid = ?, po = ?, cr = ?, record_no = ?, record_date = ?, amendment_details = ?,
                    last_modified_by = ?, last_modified_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (customer, bid, po, cr, record_no, record_date, amendment_details, username, form_id))
            
            db.execute('DELETE FROM ped_form_rows WHERE ped_form_id = ?', (form_id,))
        else:
            db.execute('''
                INSERT INTO ped_forms (po_key, customer, bid, po, cr, record_no, record_date, amendment_details, last_modified_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (po_key, customer, bid, po, cr, record_no, record_date, amendment_details, username))
            form_id = db.execute('SELECT last_insert_rowid() as id').fetchone()['id']
        
        for row in rows:
            item_no = row.get('key', '')
            if not item_no:
                continue
            
            ped_cycles_json = json.dumps(row.get('pedCycles', []))
            notes_json = json.dumps(row.get('notes', []))
            
            db.execute('''
                INSERT INTO ped_form_rows 
                (ped_form_id, item_no, part_number, part_description, rev, qty, ped_cycles, notes, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                form_id,
                item_no,
                row.get('part', ''),
                row.get('desc', ''),
                row.get('rev', ''),
                row.get('qty', ''),
                ped_cycles_json,
                notes_json,
                row.get('remarks', '')
            ))
        
        db.commit()
        db.close()
        
        return jsonify({
            'success': True,
            'lastModifiedBy': username,
            'lastModifiedAt': datetime.utcnow().isoformat()
        })
    except Exception as e:
        db.rollback()
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ped-form/load', methods=['GET'])
@login_required
def load_ped_form():
    import json
    po_key = request.args.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400
    
    db = get_db()
    try:
        cursor = db.execute('''
            SELECT id, customer, bid, po, cr, record_no, record_date, amendment_details, last_modified_by, last_modified_at
            FROM ped_forms
            WHERE po_key = ?
        ''', (po_key,))
        form = cursor.fetchone()
        
        if not form:
            db.close()
            return jsonify({'exists': False})
        
        form_id = form['id']
        
        rows_cursor = db.execute('''
            SELECT item_no, part_number, part_description, rev, qty, ped_cycles, notes, remarks
            FROM ped_form_rows
            WHERE ped_form_id = ?
            ORDER BY id
        ''', (form_id,))
        
        rows = []
        for row in rows_cursor.fetchall():
            ped_cycles = json.loads(row['ped_cycles']) if row['ped_cycles'] else []
            notes = json.loads(row['notes']) if row['notes'] else []
            rows.append({
                'key': row['item_no'],
                'part': row['part_number'] or '',
                'desc': row['part_description'] or '',
                'rev': row['rev'] or '',
                'qty': row['qty'] or '',
                'pedCycles': ped_cycles,
                'notes': notes,
                'remarks': row['remarks'] or ''
            })
        
        db.close()
        
        return jsonify({
            'exists': True,
            'customer': form['customer'] or '',
            'bid': form['bid'] or '',
            'po': form['po'] or '',
            'cr': form['cr'] or '',
            'recordNo': form['record_no'] or '',
            'recordDate': form['record_date'] or '',
            'amendmentDetails': form['amendment_details'] or '',
            'rows': rows,
            'lastModifiedBy': form['last_modified_by'] or '',
            'lastModifiedAt': form['last_modified_at'] or ''
        })
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/lead-form/save', methods=['POST'])
@login_required
def save_lead_form():
    import json
    data = request.get_json()
    
    po_key = data.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400
    
    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    record_no = data.get('recordNo', '').strip()
    record_date = data.get('recordDate', '').strip()
    rows = data.get('rows', [])
    
    username = session.get('username', 'unknown')
    
    db = get_db()
    try:
        db.execute('BEGIN TRANSACTION')
        
        cursor = db.execute('SELECT id FROM lead_forms WHERE po_key = ?', (po_key,))
        form = cursor.fetchone()
        
        if form:
            form_id = form['id']
            db.execute('''
                UPDATE lead_forms 
                SET customer = ?, bid = ?, po = ?, cr = ?, record_no = ?, record_date = ?,
                    last_modified_by = ?, last_modified_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (customer, bid, po, cr, record_no, record_date, username, form_id))
            
            db.execute('DELETE FROM lead_form_rows WHERE lead_form_id = ?', (form_id,))
        else:
            db.execute('''
                INSERT INTO lead_forms (po_key, customer, bid, po, cr, record_no, record_date, last_modified_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (po_key, customer, bid, po, cr, record_no, record_date, username))
            form_id = db.execute('SELECT last_insert_rowid() as id').fetchone()['id']
        
        for row in rows:
            item_no = row.get('itemNo', '')
            if not item_no:
                continue
            
            db.execute('''
                INSERT INTO lead_form_rows 
                (lead_form_id, item_no, part_number, part_description, rev, qty, 
                 customer_required_date, standard_lead_time, gtn_agreed_date, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                form_id,
                item_no,
                row.get('part', ''),
                row.get('desc', ''),
                row.get('rev', ''),
                row.get('qty', ''),
                row.get('customerRequiredDate', ''),
                row.get('standardLeadTime', ''),
                row.get('gtnAgreedDate', ''),
                row.get('remarks', '')
            ))
        
        db.commit()
        db.close()
        
        return jsonify({
            'success': True,
            'lastModifiedBy': username,
            'lastModifiedAt': datetime.utcnow().isoformat()
        })
    except Exception as e:
        db.rollback()
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/lead-form/load', methods=['GET'])
@login_required
def load_lead_form():
    import json
    po_key = request.args.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400
    
    db = get_db()
    try:
        cursor = db.execute('''
            SELECT id, customer, bid, po, cr, record_no, record_date, last_modified_by, last_modified_at
            FROM lead_forms
            WHERE po_key = ?
        ''', (po_key,))
        form = cursor.fetchone()
        
        if not form:
            db.close()
            return jsonify({'exists': False})
        
        form_id = form['id']
        
        rows_cursor = db.execute('''
            SELECT item_no, part_number, part_description, rev, qty, 
                   customer_required_date, standard_lead_time, gtn_agreed_date, remarks
            FROM lead_form_rows
            WHERE lead_form_id = ?
            ORDER BY id
        ''', (form_id,))
        
        rows = []
        for row in rows_cursor.fetchall():
            rows.append({
                'itemNo': row['item_no'],
                'part': row['part_number'] or '',
                'desc': row['part_description'] or '',
                'rev': row['rev'] or '',
                'qty': row['qty'] or '',
                'customerRequiredDate': row['customer_required_date'] or '',
                'standardLeadTime': row['standard_lead_time'] or '',
                'gtnAgreedDate': row['gtn_agreed_date'] or '',
                'remarks': row['remarks'] or ''
            })
        
        db.close()
        
        return jsonify({
            'exists': True,
            'customer': form['customer'] or '',
            'bid': form['bid'] or '',
            'po': form['po'] or '',
            'cr': form['cr'] or '',
            'recordNo': form['record_no'] or '',
            'recordDate': form['record_date'] or '',
            'rows': rows,
            'lastModifiedBy': form['last_modified_by'] or '',
            'lastModifiedAt': form['last_modified_at'] or ''
        })
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cr-export-excel', methods=['GET'])
@login_required
def export_cr_to_excel():
    import json
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    from openpyxl.drawing.image import Image
    from io import BytesIO
    import zipfile
    from flask import make_response
    import os
    
    def build_merged_cell_map(ws):
        merged_map = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = (min_row, min_col)
        return merged_map
    
    def write_cell(ws, row, col, value, merged_map):
        if (row, col) in merged_map:
            anchor_row, anchor_col = merged_map[(row, col)]
            cell = ws.cell(row=anchor_row, column=anchor_col)
            cell.value = value
        else:
            ws.cell(row=row, column=col, value=value)
    
    templates = {
        'CR_1': 'attached_assets/CR_1762338481711.xlsx',
        'CR_2': 'attached_assets/CR 2_1762338481710.xlsx',
        'CR_3': 'attached_assets/CR 3_1762338481711.xlsx'
    }
    
    logo_path = 'attached_assets/GTN_LOGO_1762400078631.png'
    
    for name, path in templates.items():
        if not os.path.exists(path):
            return jsonify({'error': f'Template file {name} not found'}), 404
    
    if not os.path.exists(logo_path):
        return jsonify({'error': 'GTN logo file not found'}), 404
    
    db = get_db()
    try:
        forms_cursor = db.execute('''
            SELECT id, customer, bid, po, cr, record_no, record_date, amendment_details, last_modified_by, last_modified_at
            FROM cr_forms
            ORDER BY id
        ''')
        forms = forms_cursor.fetchall()
        
        if not forms:
            db.close()
            return jsonify({'error': 'No CR forms found to export'}), 404
        
        cycle_mapping = {
            'CR_1': (0, 21),
            'CR_2': (21, 48),
            'CR_3': (48, 72)
        }
        
        remarks_column_mapping = {
            'CR_1': 27,
            'CR_2': 29,
            'CR_3': 26
        }
        
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for form_idx, form in enumerate(forms):
                form_id = form['id']
                safe_customer = ''.join(c for c in (form['customer'] or 'Customer') if c.isalnum() or c in (' ', '_', '-'))[:30]
                
                for template_key, template_path in templates.items():
                    wb = openpyxl.load_workbook(template_path)
                    ws = wb.active
                    ws.title = "CR"
                    
                    merged_map = build_merged_cell_map(ws)
                    
                    img = Image(logo_path)
                    img.width = 80
                    img.height = 60
                    ws.add_image(img, 'A1')
                    
                    if template_key == 'CR_2':
                        write_cell(ws, 1, 25, form['record_no'] or 'SAL/R02/Y', merged_map)
                        write_cell(ws, 2, 25, form['record_date'] or '', merged_map)
                    else:
                        write_cell(ws, 1, 26, form['record_no'] or 'SAL/R02/Y', merged_map)
                        write_cell(ws, 2, 26, form['record_date'] or '', merged_map)
                    
                    write_cell(ws, 3, 2, form['customer'] or '', merged_map)
                    write_cell(ws, 3, 5, form['bid'] or '', merged_map)
                    write_cell(ws, 3, 8, form['po'] or '', merged_map)
                    write_cell(ws, 3, 9, form['cr'] or '', merged_map)
                    
                    write_cell(ws, 16, 1, form['amendment_details'] or '', merged_map)
                    
                    rows_cursor = db.execute('''
                        SELECT item_no, part_number, part_description, rev, qty, cycles, remarks
                        FROM cr_form_rows
                        WHERE cr_form_id = ?
                        ORDER BY id
                    ''', (form_id,))
                    
                    data_rows = rows_cursor.fetchall()
                    data_start_row = 8
                    
                    cycle_start, cycle_end = cycle_mapping[template_key]
                    remarks_col = remarks_column_mapping[template_key]
                    
                    for row_idx, row in enumerate(data_rows):
                        excel_row = data_start_row + row_idx
                        
                        if excel_row > 12:
                            break
                        
                        write_cell(ws, excel_row, 1, row['item_no'] or '', merged_map)
                        write_cell(ws, excel_row, 2, row['part_number'] or '', merged_map)
                        write_cell(ws, excel_row, 3, row['part_description'] or '', merged_map)
                        write_cell(ws, excel_row, 4, row['rev'] or '', merged_map)
                        write_cell(ws, excel_row, 5, row['qty'] or '', merged_map)
                        
                        cycles = json.loads(row['cycles']) if row['cycles'] else []
                        
                        relevant_cycles = cycles[cycle_start:cycle_end]
                        for cycle_idx, cycle_val in enumerate(relevant_cycles):
                            col_num = 6 + cycle_idx
                            value_to_write = cycle_val if cycle_val else ''
                            write_cell(ws, excel_row, col_num, value_to_write, merged_map)
                        
                        write_cell(ws, excel_row, remarks_col, row['remarks'] or '', merged_map)
                    
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    filename = f"{template_key}_{form_idx + 1}_{safe_customer}.xlsx"
                    zip_file.writestr(filename, excel_buffer.read())
        
        db.close()
        
        zip_buffer.seek(0)
        response = make_response(zip_buffer.read())
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = 'attachment; filename=CR_Export.zip'
        
        return response
        
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
